#!/usr/bin/env python3
"""Deterministic column profiler for arbitrary Rattle ingest sources.

The FIRST step of `rattle-ingest`. Never guess a column's meaning from its
header alone: profile the VALUES (dtype, cardinality, null count, samples,
numeric stats) and combine that shape evidence with DE + EN header keywords
to produce a ranked list of candidate column roles with confidence scores.

Usage:
    profile_source.py <file> [--sheet NAME] [--json]

    <file>          .xlsx / .xlsm (openpyxl) | .csv (stdlib) | .json (stdlib,
                    list-of-dicts, e.g. the output of rattle_api.source.read_excel)
    --sheet NAME    worksheet to profile (Excel only; default: active sheet)
    --json          emit JSON to stdout (the default; accepted for explicitness)

Output: JSON to stdout, progress to stderr (house rule). Shape:

    {
      "source":   {path, filename, file_type, read_via, sheet_name, sheet_names,
                   header_row, row_count, column_count},
      "sheet_shape": {id, confidence, evidence, candidates},
      "columns": [
        {index, header, dtype, non_null, null_count, cardinality, distinct_ratio,
         samples, numeric, candidate_roles: [{role, confidence, signals}]}
      ],
      "notes": []
    }

The `columns[].candidate_roles` and `sheet_shape` feed straight into the
`source-mapping.json` contract (schemas/source-mapping.schema.json). The
profiler PROPOSES; a human confirms; nothing is written to Rattle here.

Exit codes:
    0  profiled
    1  file not found
    2  bad usage / unsupported file type / missing optional dependency

Runs without network, AI keys, or the rattle_api package. Excel reading
requires ``openpyxl`` and degrades gracefully when it is absent.
"""

from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path
from typing import Any

MAX_SAMPLES = 5
MAX_CANDIDATES = 3
LOW_CONFIDENCE = 0.60

# --------------------------------------------------------------------------
# Role keyword tables. DE first (the tenants are German), EN second.
# Matching is case-insensitive substring on the normalised header.
# Full heuristics table: ../references/column-roles.md — keep them in sync.
# --------------------------------------------------------------------------

ROLE_KEYWORDS: dict[str, list[str]] = {
    "product_name": [
        "produkt",
        "artikel",
        "artikelbezeichnung",
        "bezeichnung",
        "modell",
        "maschine",
        "typ",
        "product",
        "item",
        "model",
        "machine",
    ],
    "product_sku": [
        "artikelnummer",
        "artikelnr",
        "art.-nr",
        "art nr",
        "artnr",
        "sachnummer",
        "materialnummer",
        "matnr",
        "erp-id",
        "sku",
        "item number",
        "item no",
        "article number",
        "product code",
        "erp id",
    ],
    "area_name": [
        "bereich",
        "baugruppe",
        "sektion",
        "abschnitt",
        "zone",
        "modulgruppe",
        "area",
        "section",
        "assembly",
        "zone",
    ],
    "group_name": [
        "gruppe",
        "optionsgruppe",
        "merkmal",
        "merkmalsgruppe",
        "kategorie",
        "ausstattungsgruppe",
        "group",
        "option group",
        "feature",
        "category",
        "characteristic",
    ],
    "option_name": [
        "option",
        "variante",
        "ausführung",
        "ausfuehrung",
        "ausstattung",
        "auswahl",
        "wert",
        "variant",
        "choice",
        "value",
        "trim",
    ],
    "option_price": [
        "aufpreis",
        "mehrpreis",
        "zuschlag",
        "aufschlag",
        "optionspreis",
        "preis option",
        "surcharge",
        "option price",
        "upcharge",
        "add-on price",
        "extra",
    ],
    "base_price": [
        "grundpreis",
        "basispreis",
        "listenpreis",
        "nettopreis",
        "preis",
        "vk",
        "ek",
        "base price",
        "list price",
        "net price",
        "price",
        "msrp",
    ],
    "currency": [
        "währung",
        "waehrung",
        "currency",
        "curr",
        "iso-währung",
    ],
    "recommended_flag": [
        "serie",
        "serienmäßig",
        "serienmaessig",
        "serienausstattung",
        "standard",
        "grundausstattung",
        "basisausstattung",
        "im lieferumfang",
        "vorauswahl",
        "empfohlen",
        "default",
        "included",
        "recommended",
        "std",
    ],
    "quantity": [
        "menge",
        "stück",
        "stueck",
        "stk",
        "anzahl",
        "bedarf",
        "quantity",
        "qty",
        "count",
        "pieces",
        "amount",
    ],
    "unit": [
        "einheit",
        "me",
        "mengeneinheit",
        "uom",
        "unit",
        "unit of measure",
    ],
    "part_number": [
        "teilenummer",
        "teile-nr",
        "teilnr",
        "bauteilnummer",
        "komponentennummer",
        "part number",
        "part no",
        "part_number",
        "component number",
        "child part",
    ],
    "part_name": [
        "teilebezeichnung",
        "teilename",
        "bauteilbezeichnung",
        "benennung",
        "komponentenbezeichnung",
        "part name",
        "part_name",
        "part description",
        "component name",
        "material description",
    ],
    "parent_part_number": [
        "übergeordnet",
        "uebergeordnet",
        "oberteil",
        "hauptbaugruppe",
        "elternteil",
        "vaterartikel",
        "parent",
        "parent part",
        "assembly of",
        "next assembly",
    ],
    "bom_factor": [
        "faktor",
        "verwendungsfaktor",
        "multiplikator",
        "factor",
        "usage factor",
        "multiplier",
        "scaling",
    ],
    "number_min": ["min", "minimum", "von", "untergrenze", "number_min", "lower"],
    "number_max": ["max", "maximum", "bis", "obergrenze", "number_max", "upper"],
    "number_step": ["schritt", "schrittweite", "raster", "step", "increment"],
    "number_unit": [
        "zahleneinheit",
        "eingabeeinheit",
        "number unit",
        "number_unit",
        "input unit",
    ],
    "constraint_exclusion": [
        "ausschluss",
        "ausschließt",
        "ausschliesst",
        "nicht kombinierbar",
        "unverträglich",
        "konflikt",
        "sperrt",
        "excludes",
        "incompatible",
        "conflict",
        "forbidden",
        "not with",
    ],
    "description": [
        "beschreibung",
        "produktbeschreibung",
        "langtext",
        "bemerkung",
        "erläuterung",
        "erlaeuterung",
        "text",
        "description",
        "long text",
        "remark",
        "comment",
        "notes",
    ],
    "image_ref": [
        "bild",
        "bilddatei",
        "abbildung",
        "foto",
        "grafik",
        "image",
        "picture",
        "photo",
        "img",
        "url",
    ],
    "locale": [
        "sprache",
        "sprachcode",
        "language",
        "locale",
        "lang",
        "iso-sprache",
    ],
}

# Roles whose value shape must be numeric to be credible.
NUMERIC_ROLES = {
    "option_price",
    "base_price",
    "quantity",
    "bom_factor",
    "number_min",
    "number_max",
    "number_step",
}
# Roles whose value shape must be boolean-ish / marker-ish.
FLAG_ROLES = {"recommended_flag"}
# Roles that must be low-cardinality categorical text.
CATEGORICAL_ROLES = {"area_name", "group_name", "unit", "number_unit", "currency", "locale"}
# Roles that must be (near-)unique identifiers.
IDENTIFIER_ROLES = {"product_sku", "part_number"}

TRUE_MARKERS = {
    "x",
    "✓",
    "✔",
    "•",
    "ja",
    "j",
    "yes",
    "y",
    "true",
    "1",
    "s",
    "serie",
    "serienmäßig",
    "serienmaessig",
    "standard",
    "std",
    "inkl.",
    "inkl",
}
FALSE_MARKERS = {"", "-", "–", "—", "nein", "n", "no", "false", "0", "n/a", "na"}

CURRENCY_CODES = {"eur", "usd", "chf", "gbp", "€", "$", "£"}
UNIT_TOKENS = {
    "stk",
    "stück",
    "stueck",
    "pcs",
    "pc",
    "st",
    "m",
    "mm",
    "cm",
    "kg",
    "g",
    "l",
    "m²",
    "m2",
    "m³",
    "m3",
    "h",
    "set",
    "paar",
    "pair",
}
LOCALE_RE = re.compile(r"^[a-z]{2}([-_][A-Za-z]{2})?$", re.IGNORECASE)
IMAGE_RE = re.compile(r"(\.(png|jpe?g|webp|gif|svg)$)|^https?://", re.IGNORECASE)
DE_DECIMAL_RE = re.compile(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$|^-?\d+,\d+$")
EN_DECIMAL_RE = re.compile(r"^-?\d{1,3}(,\d{3})*(\.\d+)?$|^-?\d+\.\d+$")
CURRENCY_STRIP_RE = re.compile(r"[€$£]|\b(eur|usd|chf|gbp)\b|\s", re.IGNORECASE)
FRACTION_RE = re.compile(r"[.,]\d{1,2}$")
# Header tokeniser. Short keywords (<= SHORT_KEYWORD_LEN) match on TOKENS only —
# naive substring matching produces false friends such as "Be-ME-rkung" → unit
# and "Multi-STEP-Modul" → number_step.
TOKEN_RE = re.compile(r"[^0-9a-zà-ÿäöüß]+", re.IGNORECASE)
SHORT_KEYWORD_LEN = 4
# A "variant label" header: a concrete VALUE masquerading as a column name.
# This is the fingerprint of the wide-variant-matrix shape — the variants have
# been hoisted into the header row and the cells hold their surcharges.
VARIANT_LABEL_RE = re.compile(
    r"(\d+\s*(zoll|inch|\"|mm|cm|kw|ps|hp|l|v)\b)"
    r"|(^[A-Z]{2,}[- ]?\d)"
    r"|(\bmit\b|\bohne\b|\bwith\b|\bwithout\b)"
    r"|(\b(modul|module|paket|package|lizenz|licen[cs]e|kit|set)\b)",
    re.IGNORECASE,
)


# --------------------------------------------------------------------------
# Readers — reuse rattle_api.source semantics; do not reinvent parsing.
# --------------------------------------------------------------------------


def read_excel(path: Path, sheet: str | None) -> tuple[list[str], list[list[Any]], list[str], str]:
    """Return (headers, data_rows, sheet_names, active_sheet_name)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "Reading Excel requires openpyxl. Install: pip install openpyxl",
            file=sys.stderr,
        )
        raise SystemExit(2) from None

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    if sheet is not None:
        if sheet not in sheet_names:
            print(
                f"Sheet '{sheet}' not found. Available: {', '.join(sheet_names)}",
                file=sys.stderr,
            )
            wb.close()
            raise SystemExit(2)
        ws = wb[sheet]
    else:
        ws = wb.active
    name = str(ws.title)
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return [], [], sheet_names, name
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return headers, rows[1:], sheet_names, name


def read_csv(path: Path) -> tuple[list[str], list[list[Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect: Any = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(fh, dialect)
        rows = [list(r) for r in reader]
    if not rows:
        return [], []
    headers = [str(h).strip() for h in rows[0]]
    return headers, rows[1:]


def read_json(path: Path) -> tuple[list[str], list[list[Any]]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        print("JSON input must be a list of objects.", file=sys.stderr)
        raise SystemExit(2)
    headers: list[str] = []
    for row in data:
        if not isinstance(row, dict):
            continue
        for key in row:
            if key not in headers:
                headers.append(str(key))
    rows = [[row.get(h) for h in headers] for row in data if isinstance(row, dict)]
    return headers, rows


# --------------------------------------------------------------------------
# Value-shape profiling
# --------------------------------------------------------------------------


def _cells(rows: list[list[Any]], index: int) -> list[Any]:
    return [row[index] if index < len(row) else None for row in rows]


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _as_number(value: Any) -> float | None:
    """Parse a cell as a number, tolerating DE and EN decimal conventions."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    text = CURRENCY_STRIP_RE.sub("", value).strip()
    if not text:
        return None
    if DE_DECIMAL_RE.match(text):
        text = text.replace(".", "").replace(",", ".")
    elif EN_DECIMAL_RE.match(text):
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _looks_fractional(value: Any) -> bool:
    """True when the RAW cell was written with a decimal fraction (12,50 / 12.50)."""
    if isinstance(value, float) and not value.is_integer():
        return True
    if not isinstance(value, str):
        return False
    return bool(FRACTION_RE.search(CURRENCY_STRIP_RE.sub("", value).strip()))


def _tokens(header: str) -> set[str]:
    return {t for t in TOKEN_RE.split(header.strip().lower()) if t}


def guess_dtype(values: list[Any]) -> str:
    """Deterministic dtype guess over the non-blank values."""
    present = [v for v in values if not _is_blank(v)]
    if not present:
        return "empty"
    numeric = sum(1 for v in present if _as_number(v) is not None)
    boolish = sum(
        1
        for v in present
        if isinstance(v, bool) or str(v).strip().lower() in TRUE_MARKERS | FALSE_MARKERS
    )
    total = len(present)
    if boolish == total and numeric < total:
        return "boolean"
    if numeric == total:
        as_floats = [_as_number(v) for v in present]
        integral = all(f is not None and f.is_integer() for f in as_floats)
        # "12.000,00" is integral in value but decimal in intent — a price, not a count.
        written_fractional = any(_looks_fractional(v) for v in present)
        if integral and not written_fractional:
            return "integer"
        return "number"
    if numeric and numeric < total:
        return "mixed"
    return "string"


def profile_column(index: int, header: str, values: list[Any]) -> dict[str, Any]:
    present = [v for v in values if not _is_blank(v)]
    distinct = {str(v).strip() for v in present}
    samples = [str(v).strip()[:120] for v in list(dict.fromkeys(str(v).strip() for v in present))]
    dtype = guess_dtype(values)
    col: dict[str, Any] = {
        "index": index,
        "header": header,
        "dtype": dtype,
        "non_null": len(present),
        "null_count": len(values) - len(present),
        "cardinality": len(distinct),
        "distinct_ratio": round(len(distinct) / len(present), 3) if present else 0.0,
        "samples": samples[:MAX_SAMPLES],
    }
    numbers = [n for n in (_as_number(v) for v in present) if n is not None]
    if numbers and dtype in {"number", "integer", "mixed"}:
        col["numeric"] = {
            "min": min(numbers),
            "max": max(numbers),
            "mean": round(sum(numbers) / len(numbers), 4),
            "sum": round(sum(numbers), 4),
            "zero_count": sum(1 for n in numbers if n == 0),
        }
    return col


# --------------------------------------------------------------------------
# Role scoring: header keywords AND value shape. Neither alone is enough.
# --------------------------------------------------------------------------


def _header_score(header: str, role: str) -> tuple[float, list[str]]:
    """Score the HEADER alone. Never sufficient on its own — see _shape_score."""
    norm = header.strip().lower()
    if not norm:
        return 0.0, []
    tokens = _tokens(norm)
    best = 0.0
    signals: list[str] = []
    for kw in ROLE_KEYWORDS[role]:
        if len(kw) <= SHORT_KEYWORD_LEN and " " not in kw:
            # Short keywords ("me", "min", "step", "vk") match tokens only.
            score = 0.55 if kw in tokens else 0.0
        elif norm == kw:
            # An exact hit on a SPECIFIC keyword must outrank a suffix hit on a
            # generic one — "Vaterartikel" is a parent part, not a product.
            score = 0.70
        elif norm.startswith(kw) or norm.endswith(kw):
            score = 0.50
        elif kw in norm:
            score = 0.42
        else:
            score = 0.0
        if score > best:
            best = score
            signals = [f"header~'{kw}'"]
    return best, signals


def _shape_score(col: dict[str, Any], role: str, row_count: int) -> tuple[float, list[str]]:
    """Reward value shapes consistent with the role; punish contradictions."""
    dtype = col["dtype"]
    card = col["cardinality"]
    non_null = col["non_null"]
    ratio = col["distinct_ratio"]
    samples = [s.lower() for s in col["samples"]]
    score = 0.0
    signals: list[str] = []

    if role in NUMERIC_ROLES:
        if dtype in {"number", "integer"}:
            score += 0.28
            signals.append(f"dtype={dtype}")
        elif dtype == "mixed":
            score += 0.08
            signals.append("dtype=mixed")
        else:
            score -= 0.35
            signals.append(f"contradiction:dtype={dtype}")
        if role == "option_price":
            if col.get("numeric", {}).get("zero_count", 0) > 0:
                score += 0.05
                signals.append("has-zero-priced-rows")
            if _looks_like_variant_label(col["header"]):
                # Wide-variant-matrix fingerprint: the header IS the option name,
                # the cells are its surcharges. Never enough on its own to clear
                # the review floor — the GROUP is not stated anywhere in the sheet.
                score += 0.30
                signals.append("variant-label-header")
        if role == "base_price" and non_null and non_null == row_count:
            score += 0.04
            signals.append("dense")

    if role in FLAG_ROLES:
        if dtype == "boolean":
            score += 0.32
            signals.append("dtype=boolean")
        elif samples and all(s in TRUE_MARKERS | FALSE_MARKERS for s in samples):
            score += 0.28
            signals.append("marker-values")
        elif card <= 3 and dtype == "string":
            score += 0.10
            signals.append("low-cardinality-text")
        else:
            score -= 0.25
            signals.append(f"contradiction:dtype={dtype}")

    if role in CATEGORICAL_ROLES:
        if dtype in {"number", "integer"}:
            score -= 0.30
            signals.append("contradiction:numeric")
        elif non_null and ratio <= 0.35:
            score += 0.24
            signals.append(f"low-cardinality({card})")
        elif card <= 12:
            score += 0.12
            signals.append(f"cardinality={card}")
        else:
            score -= 0.15
            signals.append("too-many-distinct")
        if role == "currency" and samples and all(s in CURRENCY_CODES for s in samples):
            score += 0.30
            signals.append("iso-currency-values")
        if role in {"unit", "number_unit"} and samples and all(s in UNIT_TOKENS for s in samples):
            score += 0.30
            signals.append("uom-values")
        if role == "locale" and samples and all(LOCALE_RE.match(s) for s in samples):
            score += 0.30
            signals.append("locale-values")

    if role in IDENTIFIER_ROLES:
        if non_null and ratio >= 0.90:
            score += 0.26
            signals.append("near-unique")
        elif non_null and ratio >= 0.60:
            score += 0.10
            signals.append(f"distinct_ratio={ratio}")
        else:
            score -= 0.20
            signals.append("not-unique")
        if samples and all(re.search(r"\d", s) for s in samples):
            score += 0.06
            signals.append("digit-bearing")

    if role in {"product_name", "part_name"}:
        if dtype == "string" and non_null:
            score += 0.18
            signals.append("dtype=string")
        else:
            score -= 0.30
            signals.append(f"contradiction:dtype={dtype}")

    if role == "parent_part_number":
        # A parent-part column REPEATS (one parent, many children) — that is the
        # signal that separates it from a product-identity column.
        if non_null and 0.0 < ratio < 1.0:
            score += 0.16
            signals.append(f"repeating-identifier(distinct_ratio={ratio})")

    if role == "option_name":
        if dtype == "string" and non_null and ratio >= 0.30:
            score += 0.18
            signals.append("varied-text")
        elif dtype in {"number", "integer"}:
            score -= 0.30
            signals.append("contradiction:numeric")

    if role == "description":
        avg = sum(len(s) for s in col["samples"]) / len(col["samples"]) if col["samples"] else 0
        if avg >= 40:
            score += 0.24
            signals.append("long-text")
        elif dtype in {"number", "integer"}:
            score -= 0.30
            signals.append("contradiction:numeric")

    if role == "image_ref":
        if samples and all(IMAGE_RE.search(s) for s in samples):
            score += 0.34
            signals.append("image-path-or-url")
        else:
            score -= 0.25
            signals.append("no-image-shape")

    if role == "constraint_exclusion":
        if dtype == "string" and any("," in s or ";" in s for s in samples):
            score += 0.12
            signals.append("list-like")

    if role in {"part_number", "parent_part_number"} and dtype in {"number", "integer"}:
        # ERP part numbers are frequently numeric — do not punish them as text.
        score += 0.06
        signals.append("numeric-part-number")

    return score, signals


def score_roles(col: dict[str, Any], row_count: int) -> list[dict[str, Any]]:
    """Rank every role for one column. Header keywords AND value shape."""
    ranked: list[dict[str, Any]] = []
    for role in ROLE_KEYWORDS:
        h_score, h_signals = _header_score(col["header"], role)
        s_score, s_signals = _shape_score(col, role, row_count)
        if h_score == 0.0 and s_score <= 0:
            continue
        # A header hit with a contradicting shape is worse than no hit at all.
        total = h_score + s_score
        if total <= 0.05:
            continue
        ranked.append(
            {
                "role": role,
                "confidence": round(min(total, 0.99), 2),
                "signals": h_signals + s_signals,
            }
        )
    ranked.sort(key=lambda r: (-r["confidence"], r["role"]))
    return ranked[:MAX_CANDIDATES]


# --------------------------------------------------------------------------
# Sheet-shape detection — see ../references/sheet-shapes.md
# --------------------------------------------------------------------------


def _top_role(col: dict[str, Any]) -> str | None:
    cands = col.get("candidate_roles") or []
    return str(cands[0]["role"]) if cands else None


def _looks_like_variant_label(header: str) -> bool:
    return bool(header.strip()) and bool(VARIANT_LABEL_RE.search(header))


def detect_sheet_shape(columns: list[dict[str, Any]], row_count: int) -> dict[str, Any]:
    """Rank the five sheet shapes from column roles + value shape. Never the header alone."""
    roles = [_top_role(c) for c in columns]
    has = roles.count

    def _confident(col: dict[str, Any], role: str) -> bool:
        cands = col.get("candidate_roles") or []
        return bool(cands and cands[0]["role"] == role and cands[0]["confidence"] >= LOW_CONFIDENCE)

    # A variant column: numeric values under a header that is really a VALUE.
    # A column only escapes this test by being a CONFIDENT base price / quantity —
    # a weak base_price guess must not mask the matrix.
    numeric_variant_cols = [
        c
        for c in columns
        if c["dtype"] in {"number", "integer"}
        and _looks_like_variant_label(c["header"])
        and not _confident(c, "base_price")
        and not _confident(c, "quantity")
    ]
    name_cols = [c for c in columns if _top_role(c) in {"product_name", "product_sku"}]
    identity_col = next((c for c in columns if _top_role(c) == "product_name"), None)

    scores: dict[str, float] = {
        "one-row-per-product": 0.0,
        "one-row-per-option": 0.0,
        "one-row-per-bom-line": 0.0,
        "wide-variant-matrix": 0.0,
        "mixed": 0.15,
    }
    evidence: list[str] = []

    if len(numeric_variant_cols) >= 2 and name_cols:
        scores["wide-variant-matrix"] += 0.55 + min(0.20, 0.05 * len(numeric_variant_cols))
        evidence.append(
            f"{len(numeric_variant_cols)} numeric columns whose headers are variant labels "
            f"({', '.join(repr(c['header']) for c in numeric_variant_cols[:4])}) alongside "
            f"{len(name_cols)} identity column(s) → variants live in the header row"
        )

    if has("parent_part_number") and has("part_number"):
        scores["one-row-per-bom-line"] += 0.70
        evidence.append("both a parent-part and a part-number column present → parent→child edges")
    elif has("part_number") and (has("quantity") or has("bom_factor")):
        scores["one-row-per-bom-line"] += 0.45
        evidence.append("part-number column with quantity/factor → BOM lines")

    if has("option_name") and (has("group_name") or has("option_price")):
        scores["one-row-per-option"] += 0.60
        evidence.append("option-name column present with a group or option-price column")
    if has("group_name") and identity_col and identity_col["distinct_ratio"] < 0.5:
        scores["one-row-per-option"] += 0.15
        evidence.append(
            "product identity repeats across rows → rows are sub-entities, not products"
        )

    # One-row-per-product only holds when the variants are NOWHERE — neither in a
    # column (option_name) nor in the header row (numeric_variant_cols).
    if (
        identity_col is not None
        and identity_col["distinct_ratio"] >= 0.95
        and not has("option_name")
        and not numeric_variant_cols
    ):
        scores["one-row-per-product"] += 0.55
        evidence.append("product column is (near-)unique per row and no option column exists")
    if has("base_price") and row_count and not numeric_variant_cols and not has("option_name"):
        scores["one-row-per-product"] += 0.15
        evidence.append("a single base-price column and no per-variant columns")

    distinct_families = sum(
        1
        for family in (
            {"option_name", "option_price"},
            {"part_number", "parent_part_number"},
            {"product_name", "base_price"},
        )
        if any(r in family for r in roles if r)
    )
    if distinct_families >= 3:
        scores["mixed"] += 0.35
        evidence.append("product, option, and BOM column families all present in one sheet")

    ranked = sorted(scores.items(), key=lambda kv: (-kv[1], kv[0]))
    best_id, best_score = ranked[0]
    if best_score <= 0.20:
        best_id, best_score = "mixed", max(best_score, 0.30)
        evidence.append("no shape scored decisively — treat as mixed and split the sheet by hand")
    return {
        "id": best_id,
        "confidence": round(min(best_score, 0.99), 2),
        "evidence": evidence,
        "candidates": [
            {"id": sid, "confidence": round(min(sc, 0.99), 2)} for sid, sc in ranked[:3] if sc > 0
        ],
    }


# --------------------------------------------------------------------------
# Orchestration
# --------------------------------------------------------------------------


def apply_shape_context(columns: list[dict[str, Any]], shape_id: str) -> None:
    """Let the sheet shape disambiguate what a single column cannot.

    A generic name column ("Bezeichnung") is `product.name` in a product sheet and
    `part.part_name` in a BOM sheet — the column alone cannot tell you which.
    In a one-row-per-bom-line sheet, re-rank `product_name` → `part_name` and keep
    the displaced role as the runner-up so the reviewer can see what was overruled.
    `PartCreateRequest` requires BOTH part_number and part_name — without this the
    BOM path cannot emit a part at all.
    """
    if shape_id != "one-row-per-bom-line":
        return
    for col in columns:
        cands = col.get("candidate_roles") or []
        if not cands or cands[0]["role"] != "product_name":
            continue
        demoted = dict(cands[0])
        cands[0] = {
            "role": "part_name",
            "confidence": demoted["confidence"],
            "signals": [*demoted["signals"], "shape-context:one-row-per-bom-line"],
        }
        cands.insert(1, demoted)
        del cands[MAX_CANDIDATES:]


def profile(
    path: Path,
    headers: list[str],
    rows: list[list[Any]],
    file_type: str,
    read_via: str,
    sheet_name: str | None = None,
    sheet_names: list[str] | None = None,
) -> dict[str, Any]:
    row_count = len(rows)
    notes: list[str] = []

    columns: list[dict[str, Any]] = []
    for index, header in enumerate(headers):
        col = profile_column(index, header, _cells(rows, index))
        col["candidate_roles"] = score_roles(col, row_count)
        columns.append(col)

    # Shape first, then let the shape re-rank the roles it disambiguates.
    shape = detect_sheet_shape(columns, row_count)
    apply_shape_context(columns, str(shape["id"]))

    seen: dict[str, int] = {}
    for col in columns:
        index, header = col["index"], col["header"]
        top = col["candidate_roles"][0] if col["candidate_roles"] else None
        if top is None:
            notes.append(
                f"columns[{index}] '{header}': no role scored above the floor — "
                f"map it by hand or list it under unmapped_columns. Never drop it silently."
            )
        elif top["confidence"] < LOW_CONFIDENCE:
            notes.append(
                f"columns[{index}] '{header}': best role '{top['role']}' at "
                f"{top['confidence']} < {LOW_CONFIDENCE} — raise a low-confidence-mapping warning "
                f"and set review_required=true."
            )
        if col["dtype"] == "empty":
            notes.append(f"columns[{index}] '{header}': every cell is empty.")
        key = str(header).strip().lower()
        if key:
            seen[key] = seen.get(key, 0) + 1

    for key, count in seen.items():
        if count > 1:
            notes.append(
                f"duplicate header '{key}' appears {count}× — disambiguate before mapping."
            )
    if sheet_names and len(sheet_names) > 1:
        notes.append(
            f"{len(sheet_names)} worksheets present ({', '.join(sheet_names)}); "
            f"profiled '{sheet_name}'. Ingest one sheet per source-mapping."
        )
    cands = shape["candidates"]
    if len(cands) >= 2 and abs(cands[0]["confidence"] - cands[1]["confidence"]) < 0.15:
        notes.append(
            f"sheet_shape is ambiguous: '{cands[0]['id']}' ({cands[0]['confidence']}) vs "
            f"'{cands[1]['id']}' ({cands[1]['confidence']}) — raise sheet-shape-ambiguous."
        )

    return {
        "source": {
            "path": str(path),
            "filename": path.name,
            "file_type": file_type,
            "read_via": read_via,
            "sheet_name": sheet_name,
            "sheet_names": sheet_names or [],
            "header_row": 0,
            "row_count": row_count,
            "column_count": len(headers),
        },
        "sheet_shape": shape,
        "columns": columns,
        "notes": notes,
    }


def load(
    path: Path, sheet: str | None
) -> tuple[list[str], list[list[Any]], str, str, str | None, list[str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        headers, rows, sheet_names, active = read_excel(path, sheet)
        return headers, rows, suffix.lstrip("."), "openpyxl", active, sheet_names
    if suffix == ".csv":
        headers, rows = read_csv(path)
        return headers, rows, "csv", "csv", None, []
    if suffix == ".json":
        headers, rows = read_json(path)
        return headers, rows, "json", "json", None, []
    print(
        f"Unsupported file type: {suffix or '(none)'}. "
        f"Supported: .xlsx, .xlsm, .csv, .json. "
        f"For .pdf / .docx, extract text with rattle_api.source.read_source first.",
        file=sys.stderr,
    )
    raise SystemExit(2)


def main(argv: list[str]) -> int:
    args = list(argv[1:])
    if not args or "-h" in args or "--help" in args:
        print(__doc__, file=sys.stderr)
        return 0 if args else 2
    sheet: str | None = None
    if "--sheet" in args:
        i = args.index("--sheet")
        if i + 1 >= len(args):
            print("--sheet requires a worksheet name", file=sys.stderr)
            return 2
        sheet = args[i + 1]
        del args[i : i + 2]
    if "--json" in args:
        args.remove("--json")  # JSON on stdout is the default; the flag is explicit-only.
    if len(args) != 1:
        print(__doc__, file=sys.stderr)
        return 2

    path = Path(args[0])
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    headers, rows, file_type, read_via, sheet_name, sheet_names = load(path, sheet)
    if not headers:
        print(f"No header row found in {path}", file=sys.stderr)
        return 2
    print(
        f"Profiling {path.name}"
        + (f" [{sheet_name}]" if sheet_name else "")
        + f" — {len(rows)} data rows × {len(headers)} columns",
        file=sys.stderr,
    )

    result = profile(path, headers, rows, file_type, read_via, sheet_name, sheet_names)
    json.dump(result, sys.stdout, ensure_ascii=False, indent=2)
    sys.stdout.write("\n")
    print(
        f"sheet_shape={result['sheet_shape']['id']} "
        f"(confidence {result['sheet_shape']['confidence']}), "
        f"{len(result['notes'])} note(s)",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
