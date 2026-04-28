#!/usr/bin/env python3
"""Deterministic anti-pattern scanner for Rattle pricelist inputs.

Mirror of ``rattle_api.knowledge.detect_anti_patterns`` so the skill bundle
is self-contained and runnable by any AI client without depending on the
package being installed.

Usage:
    detect_anti_patterns.py <file.xlsx | file.json>

Reads a pricelist (Excel or JSON list-of-dicts) and prints a JSON array of
findings to stdout. Each finding is shaped:

    {
        "pattern_id": "implicit-base-config",
        "pattern_name": "Implicit Base Configuration",
        "row_index": 7,
        "column": "Standard",
        "value": "Standard 17-inch wheels",
        "indicator": "standard",
        "correction": "Create an explicit group with explicit options ..."
    }

Anti-pattern definitions match the four catalogued in
``skills/rattle-configurator/references/anti-patterns.md``.

Runs without network or AI keys. Excel reading requires ``openpyxl``;
JSON reading requires only the standard library.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

ANTI_PATTERNS: list[dict[str, Any]] = [
    {
        "id": "implicit-base-config",
        "name": "Implicit Base Configuration",
        "indicators": [
            "standard",
            "Grundausstattung",
            "Serienausstattung",
            "im Lieferumfang",
            "included",
            "inkl.",
            "serienmäßig",
            "Basisausstattung",
        ],
        "correction": (
            "Create an explicit group with explicit options for ALL "
            "variants — including the standard one. Mark the standard "
            "option as recommended=true."
        ),
    },
    {
        "id": "addon-only-options",
        "name": "Add-on Only Options",
        "indicators": [
            "Aufpreis",
            "Zuschlag",
            "surcharge",
            "zusätzlich",
            "extra",
            "Mehrpreis",
            "Aufschlag",
            "optional",
        ],
        "correction": (
            "For every add-on, identify the base variant it replaces or "
            "supplements. Create a group with both the base and the "
            "add-on as explicit options."
        ),
    },
    {
        "id": "description-area-smell",
        "name": "Narrative Area Smell",
        "indicators": [
            "Beschreibung",
            "Produktbeschreibung",
            "Description",
            "Overview",
            "Übersicht",
            "Mechanics",
            "Mechanik",
            "Sensorik",
            "Elektronik",
            "Bedienung",
        ],
        "correction": (
            "Narrative content does not belong in a configuration area. "
            "Create a document template (doc_type='offer') with a "
            "'Product Overview' chapter and attach a static EditorJS "
            "content block carrying the narrative."
        ),
    },
    {
        "id": "addon-only-software-modules",
        "name": "Add-on Only Software Modules",
        "indicators": [
            "Software-Modul",
            "Software Modul",
            "Modul-Aufpreis",
            "Lizenzmodul",
            "zusätzliches Modul",
            "Software surcharge",
        ],
        "correction": (
            "Create a group for the software capability with both the "
            "baseline option (price 0, recommended) and the upgrade "
            "module option."
        ),
    },
]


def detect_anti_patterns(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Scan a list of dict rows for anti-pattern indicators.

    Returns one finding per (cell, anti-pattern) hit. Uses
    case-insensitive substring matching on the indicator keywords.
    """
    detections: list[dict[str, Any]] = []
    for row_idx, row in enumerate(rows):
        for col, value in row.items():
            if value is None:
                continue
            cell_str = str(value).strip()
            if not cell_str:
                continue
            cell_lower = cell_str.lower()
            for pattern in ANTI_PATTERNS:
                for indicator in pattern["indicators"]:
                    if indicator.lower() in cell_lower:
                        detections.append(
                            {
                                "pattern_id": pattern["id"],
                                "pattern_name": pattern["name"],
                                "row_index": row_idx,
                                "column": col,
                                "value": cell_str[:200],
                                "indicator": indicator,
                                "correction": pattern["correction"],
                            }
                        )
                        break
    return detections


def read_excel(path: Path) -> list[dict[str, Any]]:
    """Read the first sheet of an .xlsx as a list of dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "Reading Excel requires openpyxl. Install: pip install openpyxl",
            file=sys.stderr,
        )
        raise SystemExit(2)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
    rows: list[dict[str, Any]] = []
    for row in rows_iter:
        rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return rows


def read_json(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise SystemExit("JSON input must be a list of objects.")
    return [dict(row) for row in data if isinstance(row, dict)]


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print(__doc__, file=sys.stderr)
        return 2
    path = Path(argv[1])
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        return 1
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        rows = read_excel(path)
    elif suffix == ".json":
        rows = read_json(path)
    else:
        print(f"Unsupported file type: {suffix}", file=sys.stderr)
        return 2
    findings = detect_anti_patterns(rows)
    json.dump(findings, sys.stdout, ensure_ascii=False, indent=2)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
