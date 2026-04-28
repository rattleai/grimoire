import os

from openpyxl import load_workbook

# source/ directory lives at the project root, one level above this package
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_DIR = os.path.join(_ROOT, "source")


def list_sources(tenant):
    tenant_dir = os.path.join(SOURCE_DIR, tenant.lower())
    if not os.path.isdir(tenant_dir):
        return []
    results = []
    for root, _, files in os.walk(tenant_dir):
        for f in files:
            if not f.startswith("."):
                results.append(os.path.relpath(os.path.join(root, f), tenant_dir))
    return sorted(results)


def read_excel(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:]]


def read_pdf(filepath):
    """Extract text from a PDF file.

    Returns:
        str: full text content of the PDF.

    Raises:
        ImportError: if pymupdf is not installed.
    """
    try:
        import fitz  # pymupdf
    except ImportError:
        raise ImportError(
            "pymupdf is required for PDF reading. "
            "Install it with: pip install 'grimoire[pdf]'"
        )
    doc = fitz.open(filepath)
    pages = []
    for page in doc:
        pages.append(page.get_text())
    doc.close()
    return "\n".join(pages)


def read_docx(filepath):
    """Extract text from a Word .docx file.

    Returns:
        str: full text content of the document.

    Raises:
        ImportError: if python-docx is not installed.
    """
    try:
        import docx
    except ImportError:
        raise ImportError(
            "python-docx is required for Word document reading. "
            "Install it with: pip install 'grimoire[docx]'"
        )
    doc = docx.Document(filepath)
    return "\n".join(p.text for p in doc.paragraphs)


def read_source(filepath):
    """Read any supported source file, dispatching by extension.

    Returns:
        dict with keys:
            type: ``"excel"`` | ``"pdf"`` | ``"docx"``
            data: ``list[dict]`` (excel) or ``str`` (pdf/docx)
            filename: basename of the file

    Raises:
        ValueError: for unsupported file extensions.
    """
    ext = os.path.splitext(filepath)[1].lower()
    filename = os.path.basename(filepath)

    if ext in (".xlsx", ".xlsm"):
        return {"type": "excel", "data": read_excel(filepath), "filename": filename}
    elif ext == ".pdf":
        return {"type": "pdf", "data": read_pdf(filepath), "filename": filename}
    elif ext == ".docx":
        return {"type": "docx", "data": read_docx(filepath), "filename": filename}
    else:
        raise ValueError(
            f"Unsupported file extension '{ext}'. Supported: .xlsx, .xlsm, .pdf, .docx"
        )
